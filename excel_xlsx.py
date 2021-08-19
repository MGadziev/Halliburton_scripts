import os
import shutil
import csv
import re
import openpyxl
import xlwt

from openpyxl import load_workbook
wb= load_workbook('test.xlsx')
sheet = wb.get_sheet_by_name('list1')
i=2

while i<=40424:
    b='A'+str(i)
    a = sheet[b].value
    if (sheet['G'+str(i)].value or sheet['H'+str(i)].value or sheet['I'+str(i)].value or sheet['J'+str(i)].value or sheet['K'+str(i)].value or sheet['L'+str(i)].value or sheet['M'+str(i)].value or sheet['N'+str(i)].value):
        book = xlwt.Workbook(encoding="utf-8")
        sheet1 = book.add_sheet("abc")
        
        sheet1.write(1, 1, sheet['F'+str(i)].value)
        sheet1.write(1, 0, a)
        sheet1.write(1, 2, sheet['G'+str(i)].value)
        sheet1.write(1, 3, sheet['H'+str(i)].value)
        sheet1.write(1, 4, sheet['I'+str(i)].value)
        sheet1.write(1, 5, sheet['J'+str(i)].value)
        sheet1.write(1, 6, sheet['K'+str(i)].value)
        sheet1.write(1, 7, sheet['L'+str(i)].value)
        sheet1.write(1, 8, sheet['M'+str(i)].value)
        sheet1.write(1, 9, sheet['N'+str(i)].value)
        sheet1.write(1, 10, sheet['B'+str(i)].value)
        sheet1.write(1, 11, sheet['E'+str(i)].value)

        sheet1.write(0, 1, 'Description(en)')
        sheet1.write(0, 0, 'Article')
        sheet1.write(0, 2, 'Description CCD (ru)')
        sheet1.write(0, 3, 'Description CCD DME (ru)')
        sheet1.write(0, 4, 'Description CCD SVO (ru)')
        sheet1.write(0, 5, 'Description CCD NOI (ru)')
        sheet1.write(0, 6, 'Description CCD PSH (ru)')
        sheet1.write(0, 7, 'Description CCD UUS (ru)')
        sheet1.write(0, 8, 'Description Inv LED (ru)')
        sheet1.write(0, 9, 'Description Inv (ru)')
        sheet1.write(0, 10, 'Country of Origin')
        sheet1.write(0, 11, 'Manufacturer')
        
        book.save(str(a)+'.xls')
    i=i+1
       

